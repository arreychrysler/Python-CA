'''
this programe edits and update and excel file from
 @helpinghands.org to @helpinghands.cm
'''

import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']  # sheet 1 is the active cell that's why it is called in the workbook(wb)
# cell = sheet['a1']

'''
#------- defining the rows and column ----#
row = sheet.max_row
column = sheet.max_column  '''

# ----- creating a new row for updated email -----#

old_domain = 'helpinghands.cm '
new_domain = 'handsinhands.org'
for i in range(2, sheet.max_row + 1):
    cell = sheet.cell(i, 3)
    if old_domain in cell.value:
        updated_email = (cell.value).replace(old_domain, new_domain)

        sheet.cell(i, 3).value = updated_email
wb.save('updated_emails.csv')

cell = sheet.cell(2, 3)  # reads the data in row 1 column 2
# cell = sheet.cell(row = 1 , column = 2)#
print(cell.value)





     
